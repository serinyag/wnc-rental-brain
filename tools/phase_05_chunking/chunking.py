from __future__ import annotations

import hashlib
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable

from docx import Document
from openpyxl import load_workbook


MAX_CHUNK_TOKENS = 450
CHUNKING_STRATEGY_CODE = "semantic_boundary_first"
CHUNKING_STRATEGY_VERSION = "pilot_v1"


@dataclass(frozen=True)
class PilotDocumentConfig:
    document_code: str
    canonical_title: str
    relative_path: str
    parser_kind: str
    parser_version: str
    top_level_include_numbers: tuple[int, ...] = ()


@dataclass(frozen=True)
class ChunkRecord:
    chunk_ordinal: int
    section_heading: str | None
    heading_path: str | None
    question_label: str | None
    document_title_snapshot: str
    body_text: str
    source_locator: str
    token_count: int
    content_hash: str
    semantic_split_decision: str
    oversized_section_handling: str


@dataclass(frozen=True)
class PilotDocumentResult:
    document_code: str
    canonical_title: str
    relative_path: str
    parser_kind: str
    parser_version: str
    chunking_strategy_code: str
    chunking_strategy_version: str
    max_chunk_tokens: int
    chunks: list[ChunkRecord]


PILOT_DOCUMENTS: tuple[PilotDocumentConfig, ...] = (
    PilotDocumentConfig(
        document_code="OPS-001",
        canonical_title="WNC Venue Rental Operations Manual",
        relative_path="sources/phase-01-03/Venue & Operations/WNC Venue Rental Operations Manual.docx",
        parser_kind="docx_heading_outline",
        parser_version="docx_heading_outline_v1",
    ),
    PilotDocumentConfig(
        document_code="TPL-006",
        canonical_title="WNC Rental Email Template Library",
        relative_path="sources/phase-01-03/Checklists + Templates/WNC Rental Email Template Library.docx",
        parser_kind="docx_template_library",
        parser_version="docx_template_library_v2",
    ),
    PilotDocumentConfig(
        document_code="TPL-007",
        canonical_title="Discovery Call Checklist",
        relative_path="sources/phase-01-03/Checklists + Templates/WNC Rental Discovery Call & Site Visit Checklist.docx",
        parser_kind="docx_checklist_sections",
        parser_version="docx_checklist_sections_v1",
        top_level_include_numbers=(1, 3),
    ),
    PilotDocumentConfig(
        document_code="SERV-001",
        canonical_title="WNC Rental Services Catalogue",
        relative_path="sources/phase-01-03/Catalogues/WNC Rental Services Catalogue.xlsm",
        parser_kind="xlsx_service_catalogue",
        parser_version="xlsx_service_catalogue_v2",
    ),
)


def repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def pilot_document_map() -> dict[str, PilotDocumentConfig]:
    return {config.document_code: config for config in PILOT_DOCUMENTS}


def generate_pilot_results() -> list[PilotDocumentResult]:
    results: list[PilotDocumentResult] = []
    for config in PILOT_DOCUMENTS:
        results.append(generate_document_result(config))
    return results


def generate_document_result(config: PilotDocumentConfig) -> PilotDocumentResult:
    path = repo_root() / config.relative_path
    if config.parser_kind == "docx_heading_outline":
        chunks = _chunk_ops_manual(path, config.canonical_title)
    elif config.parser_kind == "docx_template_library":
        chunks = _chunk_email_library(path, config.canonical_title)
    elif config.parser_kind == "docx_checklist_sections":
        chunks = _chunk_checklist(path, config.canonical_title, config.top_level_include_numbers)
    elif config.parser_kind == "xlsx_service_catalogue":
        chunks = _chunk_service_catalogue(path, config.canonical_title)
    else:
        raise ValueError(f"Unsupported parser_kind: {config.parser_kind}")

    return PilotDocumentResult(
        document_code=config.document_code,
        canonical_title=config.canonical_title,
        relative_path=config.relative_path,
        parser_kind=config.parser_kind,
        parser_version=config.parser_version,
        chunking_strategy_code=CHUNKING_STRATEGY_CODE,
        chunking_strategy_version=CHUNKING_STRATEGY_VERSION,
        max_chunk_tokens=MAX_CHUNK_TOKENS,
        chunks=chunks,
    )


def results_to_jsonable(results: Iterable[PilotDocumentResult]) -> list[dict]:
    return [asdict(result) for result in results]


def serialize_results(results: Iterable[PilotDocumentResult]) -> str:
    return json.dumps(results_to_jsonable(results), indent=2, ensure_ascii=True)


def _nonempty_paragraphs(path: Path) -> list[tuple[str, str]]:
    doc = Document(path)
    paragraphs: list[tuple[str, str]] = []
    for paragraph in doc.paragraphs:
        text = " ".join(paragraph.text.split())
        if not text:
            continue
        style_name = paragraph.style.name if paragraph.style else "Normal"
        paragraphs.append((text, style_name))
    return paragraphs


def _chunk_ops_manual(path: Path, title: str) -> list[ChunkRecord]:
    paragraphs = _nonempty_paragraphs(path)
    heading_stack: list[tuple[int, str]] = []
    current_heading: str | None = None
    current_path: str | None = None
    current_body: list[str] = []
    sections: list[tuple[str | None, str | None, list[str]]] = []
    skip_contents = False

    def flush_current() -> None:
        nonlocal current_heading, current_path, current_body
        if current_heading and current_body:
            sections.append((current_heading, current_path, current_body[:]))
        current_body = []

    for text, style_name in paragraphs:
        level = _extract_heading_level(style_name)
        if level is not None:
            if text == "Contents":
                skip_contents = True
                continue
            if skip_contents and level == 1 and text.startswith("1. "):
                skip_contents = False
            elif skip_contents:
                continue

            flush_current()
            while heading_stack and heading_stack[-1][0] >= level:
                heading_stack.pop()
            heading_stack.append((level, text))
            current_heading = text
            current_path = " > ".join(item[1] for item in heading_stack)
            continue

        if skip_contents:
            continue
        if current_heading is None:
            continue
        current_body.append(text)

    flush_current()

    chunks: list[ChunkRecord] = []
    ordinal = 1
    for heading, heading_path, body in sections:
        question_label = None
        split_sections = _split_semantic_section(
            title=title,
            section_heading=heading,
            heading_path=heading_path,
            question_label=question_label,
            paragraphs=body,
            base_source_locator=f"Heading path: {heading_path}",
            semantic_split_decision="heading/subheading block",
        )
        for chunk in split_sections:
            chunks.append(_with_ordinal(chunk, ordinal))
            ordinal += 1
    return chunks


def _chunk_email_library(path: Path, title: str) -> list[ChunkRecord]:
    paragraphs = _nonempty_paragraphs(path)
    if not paragraphs:
        return []

    intro_body: list[str] = []
    sections: list[tuple[str, list[str]]] = []
    current_heading: str | None = None
    current_body: list[str] = []

    for text, style_name in paragraphs:
        if style_name == "Heading 1":
            if current_heading is None:
                current_heading = text
                continue
            if current_body:
                sections.append((current_heading, current_body[:]))
            current_heading = text
            current_body = []
            continue

        if current_heading == title:
            intro_body.append(text)
        else:
            current_body.append(text)

    if current_heading and current_heading != title and current_body:
        sections.append((current_heading, current_body[:]))

    chunks: list[ChunkRecord] = []
    ordinal = 1

    if intro_body:
        intro_chunk = _build_chunk(
            section_heading="Library purpose and use principles",
            heading_path="Library purpose and use principles",
            question_label="How should the email template library be used?",
            document_title_snapshot=title,
            body_text="\n".join(intro_body),
            source_locator="Template library introduction",
            semantic_split_decision="intro guidance block",
            oversized_section_handling="not needed",
        )
        chunks.append(_with_ordinal(intro_chunk, ordinal))
        ordinal += 1

    for heading, body in sections:
        cleaned_heading = _strip_numeric_prefix(heading)
        split_sections = _split_semantic_section(
            title=title,
            section_heading=cleaned_heading,
            heading_path=cleaned_heading,
            question_label=f"How should {cleaned_heading.lower()} be handled?",
            paragraphs=_format_template_section(body),
            base_source_locator=f"Template heading: {heading}",
            semantic_split_decision="complete reusable template block",
        )
        for chunk in split_sections:
            chunks.append(_with_ordinal(chunk, ordinal))
            ordinal += 1
    return chunks


def _chunk_checklist(path: Path, title: str, include_top_level_numbers: tuple[int, ...]) -> list[ChunkRecord]:
    paragraphs = _nonempty_paragraphs(path)
    if not paragraphs:
        return []

    relevant_top_levels = set(include_top_level_numbers)
    current_top_level: str | None = None
    current_top_level_number: int | None = None
    current_subheading: str | None = None
    current_body: list[str] = []
    sections: list[tuple[str, str, list[str]]] = []

    def flush_current() -> None:
        nonlocal current_body
        if current_top_level and current_body:
            if current_subheading:
                heading_path = f"{current_top_level} > {current_subheading}"
                section_heading = current_subheading
            else:
                heading_path = current_top_level
                section_heading = current_top_level
            sections.append((section_heading, heading_path, current_body[:]))
        current_body = []

    for text, _style_name in paragraphs[1:]:
        top_level_number = _extract_top_level_number(text)
        if top_level_number is not None:
            flush_current()
            current_top_level_number = top_level_number
            current_top_level = text
            current_subheading = None
            continue

        if current_top_level_number not in relevant_top_levels:
            continue

        if _is_checklist_subheading(text):
            flush_current()
            current_subheading = text
            continue

        current_body.append(text)

    flush_current()

    chunks: list[ChunkRecord] = []
    ordinal = 1
    for section_heading, heading_path, body in sections:
        split_sections = _split_semantic_section(
            title=title,
            section_heading=section_heading,
            heading_path=heading_path,
            question_label=_checklist_question_label(section_heading),
            paragraphs=body,
            base_source_locator=f"Checklist section: {heading_path}",
            semantic_split_decision="checklist section or task group",
        )
        for chunk in split_sections:
            chunks.append(_with_ordinal(chunk, ordinal))
            ordinal += 1
    return chunks


def _chunk_service_catalogue(path: Path, title: str) -> list[ChunkRecord]:
    workbook = load_workbook(path, data_only=True, read_only=False)
    chunks: list[ChunkRecord] = []
    ordinal = 1

    services_sheet = workbook["Services catalogue"]
    for row_index, row in enumerate(
        services_sheet.iter_rows(min_row=4, max_row=services_sheet.max_row or 0, values_only=True),
        start=4,
    ):
        if not row or not row[0]:
            continue
        values = [_cell_text(value) for value in row[:12]]
        if values[10] and values[10].lower() != "active":
            continue
        service_code = values[0]
        values = _sanitize_service_catalogue_values(service_code, values)
        labels = [
            "Service code",
            "Display name",
            "Description",
            "Included activities",
            "Excluded activities",
            "Standard or manual pricing",
            "VAT rate",
            "Internal owner",
            "External supplier required",
            "Client approval required",
            "Active / inactive",
            "Internal notes",
        ]
        body_lines = [
            f"{label}: {value}"
            for label, value in zip(labels, values)
            if value
        ]
        display_name = values[1] or service_code
        section_heading = display_name
        heading_path = f"Services catalogue > {display_name}"
        chunk = _build_chunk(
            section_heading=section_heading,
            heading_path=heading_path,
            question_label=f"What does {display_name.lower()} include?",
            document_title_snapshot=title,
            body_text="\n".join(body_lines),
            source_locator=f'Worksheet "Services catalogue", row {row_index}, service code {service_code}',
            semantic_split_decision="service row",
            oversized_section_handling="not needed",
        )
        chunks.append(_with_ordinal(chunk, ordinal))
        ordinal += 1

    controlled_sheet = workbook["Controlled lists"]
    controlled_rows: list[str] = []
    for row_index, row in enumerate(
        controlled_sheet.iter_rows(min_row=1, max_row=controlled_sheet.max_row or 0, values_only=True),
        start=1,
    ):
        values = [_cell_text(value) for value in row if _cell_text(value)]
        if not values:
            continue
        controlled_rows.append(f"Row {row_index}: {' | '.join(values)}")

    if controlled_rows:
        chunk = _build_chunk(
            section_heading="Controlled lists",
            heading_path="Controlled lists",
            question_label="What controlled list values support the services catalogue?",
            document_title_snapshot=title,
            body_text="\n".join(controlled_rows),
            source_locator='Worksheet "Controlled lists"',
            semantic_split_decision="controlled list section",
            oversized_section_handling="not needed",
        )
        chunks.append(_with_ordinal(chunk, ordinal))

    return chunks


def _split_semantic_section(
    *,
    title: str,
    section_heading: str | None,
    heading_path: str | None,
    question_label: str | None,
    paragraphs: list[str],
    base_source_locator: str,
    semantic_split_decision: str,
) -> list[ChunkRecord]:
    if not paragraphs:
        return []

    groups: list[list[str]] = []
    current_group: list[str] = []
    current_tokens = 0
    for paragraph in paragraphs:
        paragraph_tokens = approximate_token_count(paragraph)
        if current_group and current_tokens + paragraph_tokens > MAX_CHUNK_TOKENS:
            groups.append(current_group)
            current_group = [paragraph]
            current_tokens = paragraph_tokens
        else:
            current_group.append(paragraph)
            current_tokens += paragraph_tokens
    if current_group:
        groups.append(current_group)

    multiple_groups = len(groups) > 1
    chunks: list[ChunkRecord] = []
    for index, group in enumerate(groups, start=1):
        locator = base_source_locator
        if multiple_groups:
            locator = f"{base_source_locator} [part {index}]"
        chunks.append(
            _build_chunk(
                section_heading=section_heading,
                heading_path=heading_path,
                question_label=question_label,
                document_title_snapshot=title,
                body_text="\n".join(group),
                source_locator=locator,
                semantic_split_decision=semantic_split_decision,
                oversized_section_handling="paragraph-group split" if multiple_groups else "not needed",
            )
        )
    return chunks


def _format_template_section(paragraphs: list[str]) -> list[str]:
    subject_index = next(
        (index for index, paragraph in enumerate(paragraphs) if paragraph.startswith("Subject:")),
        None,
    )
    email_index = next(
        (
            index
            for index, paragraph in enumerate(paragraphs)
            if paragraph == "Email:" or paragraph.startswith("Email:")
        ),
        None,
    )
    if subject_index is None and email_index is None:
        return paragraphs

    client_start_index = min(
        index for index in (subject_index, email_index) if index is not None
    )
    internal_guidance = paragraphs[:client_start_index]
    client_template: list[str] = []

    for paragraph in paragraphs[client_start_index:]:
        if paragraph.lower().startswith("important internal note:"):
            internal_guidance.append(paragraph)
            continue
        client_template.append(paragraph)

    formatted: list[str] = ["INTERNAL GUIDANCE"]
    formatted.extend(internal_guidance)
    formatted.append("")
    formatted.append("CLIENT-FACING TEMPLATE")
    formatted.extend(client_template)
    return formatted


def _sanitize_service_catalogue_values(service_code: str, values: list[str]) -> list[str]:
    sanitized = values[:]
    if service_code != "facilitator_sourcing":
        return sanitized

    internal_notes = sanitized[11]
    if not internal_notes:
        return sanitized

    filtered_lines = [
        line
        for line in internal_notes.splitlines()
        if "Facilitators & Rental Experiences catalogue" not in line
    ]
    sanitized[11] = "\n".join(line for line in filtered_lines if line.strip())
    return sanitized


def _with_ordinal(chunk: ChunkRecord, ordinal: int) -> ChunkRecord:
    return ChunkRecord(
        chunk_ordinal=ordinal,
        section_heading=chunk.section_heading,
        heading_path=chunk.heading_path,
        question_label=chunk.question_label,
        document_title_snapshot=chunk.document_title_snapshot,
        body_text=chunk.body_text,
        source_locator=chunk.source_locator,
        token_count=chunk.token_count,
        content_hash=chunk.content_hash,
        semantic_split_decision=chunk.semantic_split_decision,
        oversized_section_handling=chunk.oversized_section_handling,
    )


def _build_chunk(
    *,
    section_heading: str | None,
    heading_path: str | None,
    question_label: str | None,
    document_title_snapshot: str,
    body_text: str,
    source_locator: str,
    semantic_split_decision: str,
    oversized_section_handling: str,
) -> ChunkRecord:
    normalized_body = body_text.strip()
    content_hash = _content_hash(
        document_title_snapshot=document_title_snapshot,
        section_heading=section_heading,
        heading_path=heading_path,
        question_label=question_label,
        body_text=normalized_body,
    )
    return ChunkRecord(
        chunk_ordinal=0,
        section_heading=section_heading,
        heading_path=heading_path,
        question_label=question_label,
        document_title_snapshot=document_title_snapshot,
        body_text=normalized_body,
        source_locator=source_locator,
        token_count=approximate_token_count(normalized_body),
        content_hash=content_hash,
        semantic_split_decision=semantic_split_decision,
        oversized_section_handling=oversized_section_handling,
    )


def approximate_token_count(text: str) -> int:
    return len(re.findall(r"\w+|[^\w\s]", text, flags=re.UNICODE))


def _content_hash(
    *,
    document_title_snapshot: str,
    section_heading: str | None,
    heading_path: str | None,
    question_label: str | None,
    body_text: str,
) -> str:
    payload = "\n\n".join(
        [
            document_title_snapshot,
            section_heading or "",
            heading_path or "",
            question_label or "",
            body_text,
        ]
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _extract_heading_level(style_name: str) -> int | None:
    match = re.fullmatch(r"Heading (\d+)", style_name)
    if not match:
        return None
    return int(match.group(1))


def _extract_top_level_number(text: str) -> int | None:
    match = re.match(r"^(\d+)\.\s", text)
    if not match:
        return None
    return int(match.group(1))


def _is_checklist_subheading(text: str) -> bool:
    if text.startswith("[") or text.startswith("☐") or text.startswith("Purpose:") or text.startswith("Use this section"):
        return False
    if text.endswith(":"):
        return False
    if _extract_top_level_number(text) is not None:
        return False
    words = text.split()
    if len(words) > 6:
        return False
    if not re.search(r"[A-Za-z]", text):
        return False
    if re.search(r"[.!?]$", text):
        return False
    return text == text.title() or "&" in text or "/" in text


def _checklist_question_label(section_heading: str) -> str | None:
    cleaned = _strip_numeric_prefix(section_heading).strip()
    lower = cleaned.lower()
    if lower.startswith("space"):
        return "What space and layout details should be confirmed?"
    if lower.startswith("food"):
        return "What food, beverage, and experience details should be confirmed?"
    if lower.startswith("production"):
        return "What production, technical, and branding details should be confirmed?"
    if lower.startswith("commercial"):
        return "What commercial and decision-process details should be confirmed?"
    if lower.startswith("1. discovery"):
        return "What should be confirmed during the discovery call?"
    if lower.startswith("3. decisions"):
        return "What follow-up and next actions should be captured after the call?"
    return f"What should be confirmed about {lower}?"


def _strip_numeric_prefix(text: str) -> str:
    return re.sub(r"^\d+\.\s*", "", text).strip()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
